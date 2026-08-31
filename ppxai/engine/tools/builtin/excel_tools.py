"""
Excel tools for multimodal attachments.

Phase 4.1 (v1.17.4). Two tools that let the model explore and read Excel
files the user has attached via `/attach sheet.xlsx`:

    list_excel_sheets(file_id)
        → sheet names, row/column dimensions, data preview

    read_excel_sheet(file_id, sheet, rows, as_markdown)
        → sheet data as a markdown table or CSV text

Both tools resolve `file_id` through the engine's SessionFileStore, same
pattern as `pdf_tools.py`. Guarded by `try: import openpyxl` at
registration time — installs without the `[data]` extras group silently
skip these tools.

Chart rendering (RenderExcelChartTool via matplotlib) is deferred to a
follow-up step — it requires matplotlib which is a heavy dependency many
users won't need.
"""

from __future__ import annotations

from typing import Any

from ...file_ref import FILE_REF_PROPERTIES, resolve_file_reference
from ...types import ToolEngineProtocol, ToolManagerProtocol
from ..base import BaseTool

# Default row limits to prevent token-budget blowout on large sheets.
_DEFAULT_MAX_ROWS = 100
_MAX_ALLOWED_ROWS = 5000
_MAX_TEXT_CHARS = 100_000


def _resolve_file(
    engine: Any,
    file_id: str | None = None,
    path: str | None = None,
) -> tuple[Any | None, str | None]:
    """Resolve a file reference via the unified engine resolver.

    Accepts EITHER `file_id` (SessionFileStore chat attachment) or
    `path` (workspace file). v1.18.7 — see `engine.file_ref`.
    """
    return resolve_file_reference(engine, file_id=file_id, path=path)


def _is_excel(meta: Any) -> bool:
    """Check if a file is an Excel spreadsheet by MIME type or extension."""
    if "spreadsheet" in (meta.media_type or ""):
        return True
    name = (meta.name or "").lower()
    return name.endswith((".xlsx", ".xls", ".xlsm", ".xlsb"))


class ListExcelSheetsTool(BaseTool):
    """List all sheets in an attached Excel file with dimensions and preview."""

    def __init__(self, engine: ToolEngineProtocol):
        self.engine = engine
        self.name = "list_excel_sheets"
        self.description = (
            "List all sheets in an Excel file. Returns sheet names, row/column "
            "counts, and a preview of column headers. Use this first to "
            "understand the structure before reading specific sheets. Pass "
            "either 'file_id' (chat attachment) or 'path' (workspace file) — "
            "exactly one is required."
        )
        self.parameters = {
            "type": "object",
            "properties": dict(FILE_REF_PROPERTIES),
            "required": [],
        }

    async def execute(
        self,
        file_id: str | None = None,
        path: str | None = None,
        **kwargs,
    ) -> str:
        meta, err = _resolve_file(self.engine, file_id=file_id, path=path)
        if err:
            return f"Error: {err}"
        if not _is_excel(meta):
            return f"Error: {meta.name!r} is not an Excel file (type={meta.media_type!r})."

        try:
            import openpyxl
        except ImportError:
            return "Error: openpyxl not installed. Install with: pip install 'ppxai[data]'"

        try:
            wb = openpyxl.load_workbook(str(meta.path), read_only=True, data_only=True)
        except Exception as exc:
            return f"Error opening {meta.name!r}: {exc}"

        lines: list[str] = [f"# {meta.name} — {len(wb.sheetnames)} sheet(s)\n"]

        for i, sheet_name in enumerate(wb.sheetnames, 1):
            ws = wb[sheet_name]
            max_row = ws.max_row or 0
            max_col = ws.max_column or 0

            # Preview: first row as column headers
            headers: list[str] = []
            if max_row >= 1:
                for cell in ws[1]:
                    val = cell.value
                    headers.append(str(val) if val is not None else "")

            header_preview = " | ".join(headers[:10]) if headers else "(empty)"
            if len(headers) > 10:
                header_preview += f" ... (+{len(headers) - 10} cols)"

            lines.append(
                f"## {i}. {sheet_name}\n"
                f"- Rows: {max_row}\n"
                f"- Columns: {max_col}\n"
                f"- Headers: {header_preview}\n"
            )

        try:
            wb.close()
        except Exception:
            pass

        return "\n".join(lines)


class ReadExcelSheetTool(BaseTool):
    """Read data from a specific sheet as a markdown table or CSV."""

    def __init__(self, engine: ToolEngineProtocol):
        self.engine = engine
        self.name = "read_excel_sheet"
        self.description = (
            "Read data from a specific sheet of an Excel file. Returns the "
            "data as a markdown table by default, or as CSV. Use "
            "list_excel_sheets first to see available sheets and their "
            "dimensions. Large sheets are automatically truncated. Pass "
            "either 'file_id' (chat attachment) or 'path' (workspace file) — "
            "exactly one is required."
        )
        self.parameters = {
            "type": "object",
            "properties": {
                **FILE_REF_PROPERTIES,
                "sheet": {
                    "type": "string",
                    "description": (
                        "Sheet name to read. Use the exact name from "
                        "list_excel_sheets output."
                    ),
                },
                "rows": {
                    "type": "integer",
                    "description": (
                        f"Maximum rows to return (default {_DEFAULT_MAX_ROWS}, "
                        f"max {_MAX_ALLOWED_ROWS}). Use to limit output on "
                        f"large sheets."
                    ),
                    "default": _DEFAULT_MAX_ROWS,
                },
                "as_markdown": {
                    "type": "boolean",
                    "description": (
                        "If true (default), return as a markdown table. "
                        "If false, return as CSV text."
                    ),
                    "default": True,
                },
            },
            "required": ["sheet"],
        }

    async def execute(
        self,
        sheet: str = "",
        file_id: str | None = None,
        path: str | None = None,
        rows: int = _DEFAULT_MAX_ROWS,
        as_markdown: bool = True,
        **kwargs,
    ) -> str:
        meta, err = _resolve_file(self.engine, file_id=file_id, path=path)
        if err:
            return f"Error: {err}"
        if not _is_excel(meta):
            return f"Error: {meta.name!r} is not an Excel file."

        try:
            import openpyxl
        except ImportError:
            return "Error: openpyxl not installed. Install with: pip install 'ppxai[data]'"

        try:
            wb = openpyxl.load_workbook(str(meta.path), read_only=True, data_only=True)
        except Exception as exc:
            return f"Error opening {meta.name!r}: {exc}"

        if sheet not in wb.sheetnames:
            available = ", ".join(wb.sheetnames)
            try:
                wb.close()
            except Exception:
                pass
            return f"Error: sheet {sheet!r} not found. Available: {available}"

        try:
            max_rows = max(1, min(int(rows), _MAX_ALLOWED_ROWS))
        except (TypeError, ValueError):
            max_rows = _DEFAULT_MAX_ROWS

        ws = wb[sheet]
        total_rows = ws.max_row or 0
        total_cols = ws.max_column or 0

        if total_rows == 0:
            try:
                wb.close()
            except Exception:
                pass
            return f"{meta.name} / {sheet}: empty sheet (0 rows)."

        # Read rows up to the limit
        all_rows: list[list[str]] = []
        total_chars = 0
        truncated = False

        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i >= max_rows + 1:  # +1 to include header row
                truncated = True
                break
            cells = [str(cell) if cell is not None else "" for cell in row]
            row_text = " | ".join(cells)
            if total_chars + len(row_text) > _MAX_TEXT_CHARS:
                truncated = True
                break
            all_rows.append(cells)
            total_chars += len(row_text)

        try:
            wb.close()
        except Exception:
            pass

        if not all_rows:
            return f"{meta.name} / {sheet}: no data rows."

        # Format output
        header = f"# {meta.name} / {sheet} ({total_rows} rows × {total_cols} cols)\n\n"

        if as_markdown:
            result = header + _to_markdown_table(all_rows)
        else:
            result = header + _to_csv(all_rows)

        if truncated:
            shown = len(all_rows) - 1  # subtract header row
            result += (
                f"\n\n[Showing {shown} of {total_rows} rows. "
                f"Use rows=<N> to see more.]"
            )

        return result


def _to_markdown_table(rows: list[list[str]]) -> str:
    """Convert a list of rows (first is header) to a markdown table."""
    if not rows:
        return ""
    if len(rows) == 1:
        # Only header, no data
        header = " | ".join(rows[0])
        separator = " | ".join("---" for _ in rows[0])
        return f"| {header} |\n| {separator} |"

    header = rows[0]
    data = rows[1:]

    lines = []
    lines.append("| " + " | ".join(header) + " |")
    lines.append("| " + " | ".join("---" for _ in header) + " |")
    for row in data:
        # Pad row to match header length
        padded = row + [""] * (len(header) - len(row))
        lines.append("| " + " | ".join(padded[:len(header)]) + " |")

    return "\n".join(lines)


def _to_csv(rows: list[list[str]]) -> str:
    """Convert rows to CSV text with proper quoting."""
    lines = []
    for row in rows:
        cells = []
        for cell in row:
            if "," in cell or '"' in cell or "\n" in cell:
                cells.append('"' + cell.replace('"', '""') + '"')
            else:
                cells.append(cell)
        lines.append(",".join(cells))
    return "\n".join(lines)


# =============================================================================
# Registration
# =============================================================================


def register_tools(manager: ToolManagerProtocol, engine: ToolEngineProtocol) -> bool:
    """Register Excel tools. Returns False if openpyxl is not installed."""
    try:
        import openpyxl  # noqa: F401
    except ImportError:
        return False

    if engine is None:
        return False

    manager.register_tool(ListExcelSheetsTool(engine))
    manager.register_tool(ReadExcelSheetTool(engine))
    return True


__all__ = [
    "ListExcelSheetsTool",
    "ReadExcelSheetTool",
    "register_tools",
]
